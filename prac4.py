import firebase_admin
from firebase_admin import firestore
from firebase_admin import credentials
import openpyxl

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)

db = firestore.client()

try:
    wb = openpyxl.load_workbook("empleados.xlsx")
    hoja = wb.active
except:
    wb = openpyxl.Workbook()
    hoja = wb.active

dato = db.collection("lista").document("datos")
""" dato.set({'jugadores':[
    {'nombre':'pedro','posicion':'delantero'},
    {'nombre':'carlos','posicion':'defensa'}
]}) """
hoja["A1"] = "Nombres"
hoja["B1"] = "Posiciones"

datosRef = db.collection("lista").get()

nombres = []
posiciones = []

for admin in datosRef:
    adminDict = admin.to_dict()
    for datosl in adminDict.get('jugadores', []):
        lisNombres = datosl['nombre']
        nombres.append(lisNombres)
        lisPosiciones = datosl['posicion']
        posiciones.append(lisPosiciones)

print(nombres)
print(posiciones)
linea = hoja.max_row + 1

for nombre, posicion in zip(nombres, posiciones):
    hoja.append([nombre, posicion])
wb.save("empleados.xlsx")
print("listo")





