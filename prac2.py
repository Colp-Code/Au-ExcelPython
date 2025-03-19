import openpyxl

try:
    wb = openpyxl.load_workbook("empleados.xlsx")
    hoja = wb.active
except:
    wb = openpyxl.Workbook()
    hoja = wb.active

lina = hoja.max_row 

for i in range(lina):
    hoja.append(["Hola"])
    print(i)
    print("listo")

wb.save("empleados.xlsx")
print("Todo listo")