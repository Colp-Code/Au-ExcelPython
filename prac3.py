import firebase_admin
from firebase_admin import firestore
from firebase_admin import credentials
import openpyxl
import datetime as dt
import sys

cred = credentials.Certificate("serviceAccountKey.json")
firebase_admin.initialize_app(cred)

db = firestore.client()

try:
    wb = openpyxl.load_workbook("empleados.xlsx")
    hoja = wb.active
except:
    wb = openpyxl.Workbook()
    hoja = wb.active

linea = hoja.max_row + 1
hoja["A1"] = "Fecha"
hoja["B1"] = "Descripcion"
hoja["C1"] = "Responsable"
hoja["D1"] = "Peso"
hoja["E1"] = "Valor"
hoja["F1"] = "Tipo de salida"

current_date = dt.date.today()
format_date = current_date.strftime("%d/%m/%y")
ejemplo_data = current_date.strftime("%d%m%y")

infoDescripcion = ["Aguacate hass","Aguacate papelito","Bold","Banano","Durazno","Limion Comun","Limon mandarino","Limon tahiti","Mandarina","Mango tomy","Maracuya","Naranja tangelo","Naranja valencia","Panela","Platano","Platano maduro","Guanabana","Almuerzo-Desayuno","Arriendo casa","Arriedo local","Arriendo parqueadero","BCS carro","Chevyplan","Entrada abastos","Gasolina","Gasto personal","Insumo","Lavado","Mario","Onces","Prestamo","Parqueadero carro","Parqueadero moto","Self security","Servicio de energia","Servicio de agua","Venta caja","Venta nequi","Venta daviplata"]
infoNombres = ["Andres","Armando","Camilo P","Chalo","Dairo","Emilio","Hugo","Jairo","Jeisson","Jimmy","Jose Angel","Julian","Mario","Mayerly","Nacho","Robin","Sara","Sra Luz","Venta","Yudi"]
tipoSalida = ["Abono","Adelanto de sueldo","Arriendo","Caja menor","Deuda","Flete","Gasto alimentacion","Gasto carro","Gasto moto","Insumo","Inversion","Servicion publicos","Sueldo","Venta"]
while True:
    try:
        respuesta = int(input(f"""¿Quieres la fecha automatica o la quieres poner?
                                                    
    Poner = 1
    automatica = 2
    Cual es tu opcion: """))
        if respuesta == 1:
            fecha = str(input(f"""                       
Acuerda que el formato tiene que ser dia/mes/año
    Ejemplo {ejemplo_data}
Fecha del registro: """))
            errorTexto = int(fecha)
            fechaFinal = "/".join([fecha[i:i+2] for i in range(0,len(fecha),2)])
            print(fechaFinal[0:8])
            break
        else:
            fecha = format_date
            print(fecha)
            break
    except:
        print("Tipo de dato incorrecto intentalo de nuevo :)")
        sys.exit(1)

while True:
    try:
        print(f"""
                ===================
                Que tipo de descripcion desea?
                ===================
                """)
        for i,x in enumerate(infoDescripcion):
            print(i + 1 ,x)

        descripcion = int(input("Que opcion escojes: "))
        indexDescripcion = infoDescripcion[descripcion - 1]
        print(indexDescripcion)
        break               
    except:
        print("Algo salio mal")
        sys.exit(1)

while True:
    try:
        print(f"""
                ===================
                Quien fue el responsable?
                ===================
                """)
        for i, x in enumerate(infoNombres):
            print(i + 1 ,x)

        responsable = int(input("Responsable de este registro: "))
        indexNombre = infoNombres[responsable - 1]
        print(indexNombre)
        break
    except:
            print("Algo salio mal")
            sys.exit(1)
while True:
    try:
        print("""
                ===================
                Peso
                ===================
                """)
        peso = float(input("Peso del producto(Si no tiene dale al enter): "))
        break
    except:
        peso = 0
        print(f'El peso se escribira como 0')
        break
while True:
    try:
        print("""
                ===================
                Valor
                ===================
                """)
        valor = float(input("Valor del producto(Si no tiene dale al enter): "))
        break
    except:
        valor = 0
        print(f'El valor sera escrito como 0')
        break
while True:
    try:
        print(f"""
                ===================
                ¿Que tipo de salida es?
                ===================
                """)
        for i, x in enumerate(tipoSalida):
            print(i + 1 ,x)
        
        tipoDeSalida = int(input(f'¿El tipo de salida es: '))
        indexTipoSalida = tipoSalida[tipoDeSalida - 1]
        print(indexTipoSalida)
        break
    except:
        print("Algo salio mal")
        sys.exit(1)

print(f"""Los tados agregados fueron
La fecha es: {fecha}
La descripcion es: {indexDescripcion}
El responsable es: {indexNombre}
El peso es: {peso}
El valor es: {valor}
La fuente es: {indexTipoSalida}
""")
while True:
    try:
        confirmar = int(input("""
        ¿Quieres confirmar esta informacion?
        si = 1
        no = 2
        """))
        if confirmar not in range(1,3):
            print("Opcion no valida")
        if confirmar == 1:
            hoja.append([fecha, indexDescripcion, indexNombre, peso, valor, indexTipoSalida,])
            print("Datos listos")
            break
        elif confirmar == 2:
            print("Datos no guardados")
            break         
    except:
        print("Tipo de dato no valido")
        sys.exit(1)

wb.save("empleados.xlsx")
print("listo")