import openpyxl


nombre = str(input("Cual es tu nombre: "))
edad = int(input("Cual es tu edad: "))

wb = openpyxl.Workbook()
sheet = wb.active
print(f'Hoja activa: {sheet.title}')
sheet.title = "MiHoja"
print(f'Hoja activa: {wb.active.title}')

sheet["A1"] = "Nombre"
sheet["B1"] = "Edad"
sheet["A2"] = nombre
sheet["B2"] = edad

# Añade la hoja 'Hoja' en la primera posición. Como el nombre
# 'Hoja' ya existe, le añade el número 1 al final del nombre
hoja2 = wb.create_sheet("Hoja", 0)
print(f'La hoja {hoja2} esta crada.')

wb.save("datos.xlsx")
print("Archivo de excel creado con exito")

"""
# Añade la hoja 'Hoja' al final (por defecto)
>>> hoja1 = wb.create_sheet("Hoja")
# Añade la hoja 'Hoja' en la primera posición. Como el nombre
# 'Hoja' ya existe, le añade el número 1 al final del nombre
>>> hoja2 = wb.create_sheet("Hoja", 0)
# Añade la hoja 'Otra hoja' en la posición 1
>>> wb.create_sheet(index=1, title="Otra hoja")
# Muestra los nombres de las hojas
>>> print(wb.sheetnames)
['Hoja1', 'Otra hoja', 'Valores', 'Hoja']
"""